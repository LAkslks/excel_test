import openpyxl
import json

# Считываем данные из Excel
book = openpyxl.load_workbook('price.xlsx')


data = []


# Указываем названия листов
sheets_names_mapping = {
    "Автостекло. Аксессуары. Клей": "Иномарки",
    "Российский автопром": "Отечественные"
}


# Итерируемся по указанным листам
for sheet_name, catalog in sheets_names_mapping.items():
    if sheet_name in book.sheetnames:
        sheet = book[sheet_name]
        
        # Пропускаем заголовок и начинаем с 2-й строки
        for row in range(8, 12+ 1):
            item = {
                'catalog': catalog,  # Иномарки или Отечественные
                'category': str(sheet.cell(row=row, column=1).value),  # Вид стекла
                'art': str(sheet.cell(row=row, column=2).value),  # Код AGC
                'oldcode': str(sheet.cell(row=row, column=3).value),  # Старый Код AGC
                'eurocode': str(sheet.cell(row=row, column=4).value),  # Еврокод
                'name': str(sheet.cell(row=row, column=5).value),  # Наименование
                'price': str(sheet.cell(row=row,  column=9).value)  # ОПТ
            }

            # Проверка для фиксированной цены
            if item['price'] == '*':
                item['price'] = 'Фиксированная цена'

            data.append(item)

        

# Запись данных в файл JSON
with open('output.json', 'w', encoding='utf-8') as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print("Данные успешно записаны в output.json")